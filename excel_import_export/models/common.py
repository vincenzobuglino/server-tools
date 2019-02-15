# -*- coding: utf-8 -*-
import re
import uuid
import xlrd
import csv
import base64
import string
import itertools
from datetime import datetime as dt
from ast import literal_eval
from openpyxl.styles import colors, PatternFill, Alignment, Font
from dateutil.parser import parse
from io import StringIO
from odoo.exceptions import ValidationError
from odoo import _


def adjust_cell_formula(value, k):
    """ Cell formula, i.e., if i=5, val=?(A11)+?(B12) -> val=A16+B17  """
    if isinstance(value, str):
        for i in range(value.count('?(')):
            if value and '?(' in value and ')' in value:
                i = value.index('?(')
                j = value.index(')', i)
                val = value[i + 2:j]
                col, row = split_row_col(val)
                new_val = '%s%s' % (col, row+k)
                value = value.replace('?(%s)' % val, new_val)
    return value


def get_field_aggregation(field):
    """ i..e, 'field@{sum/average/max/min}' """
    if field and '@{' in field and '}' in field:
        i = field.index('@{')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field[:i], cond)
        except Exception:
            return (field.replace('@{%s}' % cond, ''), False)
    return (field, False)


def get_field_condition(field):
    """ i..e, 'field${value > 0 and value or False}' """
    if field and '${' in field and '}' in field:
        i = field.index('${')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field.replace('${%s}' % cond, ''), cond)
        except Exception:
            return (field, False)
    return (field, False)


def get_field_format(field):
    """
        Available formats
        - font = bold, bold_red
        - fill = red, blue, yellow, green, grey
        - align = left, center, right
        - number = true, false
        i.e., 'field#{font=bold;fill=red;align=center;number_format=number}'
    """
    if field and '#{' in field and '}' in field:
        i = field.index('#{')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field.replace('#{%s}' % cond, ''), cond)
        except Exception:
            return (field, False)
    return (field, False)


def fill_cell_format(field, field_format):
    avail_format = {
        'font': {
            'bold': Font(name="Arial", size=10, bold=True),
            'bold_red':
                Font(name="Arial", size=10, color=colors.RED, bold=True),
        },
        'fill': {
            'red': PatternFill("solid", fgColor="FF0000"),
            'grey': PatternFill("solid", fgColor="DDDDDD"),
            'yellow': PatternFill("solid", fgColor="FFFCB7"),
            'blue': PatternFill("solid", fgColor="9BF3FF"),
            'green': PatternFill("solid", fgColor="B0FF99"),
        },
        'align': {
            'left': Alignment(horizontal='left'),
            'center': Alignment(horizontal='center'),
            'right': Alignment(horizontal='right'),
        },
        'number_format': {
            'number': '#,##0.00',
            'date': 'dd/mm/yyyy',
            'datestamp': 'yyyy-mm-dd',
            'text': '@',
            'percent': '0.00%',
        },
    }
    formats = field_format.split(';')
    for f in formats:
        (key, value) = f.split('=')
        if key not in avail_format.keys():
            raise ValidationError(_('Invalid format type %s' % key))
        if value.lower() not in avail_format[key].keys():
            raise ValidationError(
                _('Invalid value %s for format type %s' % (value, key)))
        cell_format = avail_format[key][value]
        if key == 'font':
            field.font = cell_format
        if key == 'fill':
            field.fill = cell_format
        if key == 'align':
            field.alignment = cell_format
        if key == 'number_format':
            if value == 'text':
                try:
                    # In case value can't be encoded as utf, we do normal str()
                    field.value = field.value.encode('utf-8')
                except Exception:
                    field.value = str(field.value)
            field.number_format = cell_format


def get_line_max(line_field):
    """ i.e., line_field = line_ids[100], max = 100 else 0 """
    if line_field and '[' in line_field and ']' in line_field:
        i = line_field.index('[')
        j = line_field.index(']')
        max_str = line_field[i + 1:j]
        try:
            if len(max_str) > 0:
                return (line_field[:i], int(max_str))
            else:
                return (line_field, False)
        except Exception:
            return (line_field, False)
    return (line_field, False)


def get_groupby(line_field):
    """i.e., line_field = line_ids["a_id, b_id"], groupby = ["a_id", "b_id"]"""
    if line_field and '[' in line_field and ']' in line_field:
        i = line_field.index('[')
        j = line_field.index(']')
        groupby = literal_eval(line_field[i:j+1])
        return groupby
    return False


def split_row_col(pos):
    match = re.match(r"([a-z]+)([0-9]+)", pos, re.I)
    if not match:
        raise ValidationError(_('Position %s is not valid') % pos)
    col, row = match.groups()
    return col, int(row)


def openpyxl_get_sheet_by_name(book, name):
    """ Get sheet by name for openpyxl """
    i = 0
    for sheetname in book.sheetnames:
        if sheetname == name:
            return book.worksheets[i]
        i += 1
    raise ValidationError(_("'%s' sheet not found") % (name,))


def xlrd_get_sheet_by_name(book, name):
    try:
        for idx in itertools.count():
            sheet = book.sheet_by_index(idx)
            if sheet.name == name:
                return sheet
    except IndexError:
        raise ValidationError(_("'%s' sheet not found") % (name,))


def add_row_skips(vals):
    """ Add row skips for side column when match with \\skiprow """
    skips = {}
    # Find all skips requried for each rows
    for k, v in vals.items():
        if not k or '\\skiprow' not in k:
            continue  # No skip for this field
        row = 0
        for value in v:
            num_skip = isinstance(value, str) and \
                value.count('\\skiprow') or 0
            if skips.get(row, 0) < num_skip:
                skips.update({row: num_skip})
            row += 1
    if not skips:
        return vals
    # Add skips for all other fields
    for k, v in vals.items():
        row = 0
        for value in v:
            num_skip = isinstance(value, str) and \
                value.count('\\skiprow') or 0
            if skips.get(row, 0) > num_skip:
                value = str(value)
                for x in range(skips[row] - num_skip):
                    value += '\\skiprow'
                v[row] = value
            row += 1
    return vals


def isfloat(input):
    try:
        float(input)
        return True
    except ValueError:
        return False


def isinteger(input):
    try:
        int(input)
        return True
    except ValueError:
        return False


def isdatetime(input):
    try:
        if len(input) == 10:
            dt.strptime(input, '%Y-%m-%d')
        elif len(input) == 19:
            dt.strptime(input, '%Y-%m-%d %H:%M:%S')
        else:
            return False
        return True
    except ValueError:
        return False


def str_to_number(input):
    if isinstance(input, str):
        if ' ' not in input:
            if isdatetime(input):
                return parse(input)
            elif isinteger(input):
                if not (len(input) > 1 and input[:1] == '0'):
                    return int(input)
            elif isfloat(input):
                if not (input.find(".") > 2 and input[:1] == '0'):  # 00.123
                    return float(input)
    return input


def csv_from_excel(excel_content, delimiter, quote):
    decoded_data = base64.decodestring(excel_content)
    wb = xlrd.open_workbook(file_contents=decoded_data)
    sh = wb.sheet_by_index(0)
    content = StringIO()
    quoting = csv.QUOTE_ALL
    if not quote:
        quoting = csv.QUOTE_NONE
    if delimiter == " " and quoting == csv.QUOTE_NONE:
        quoting = csv.QUOTE_MINIMAL
    wr = csv.writer(content, delimiter=delimiter, quoting=quoting)
    for rownum in xrange(sh.nrows):
        row = []
        for x in sh.row_values(rownum):
            # if isinstance(x, str):
            #     x = x.strip()
            if quoting == csv.QUOTE_NONE and delimiter in x:
                raise ValidationError(
                    _('Template with CSV Quoting = False, data must not '
                      'containg same char with delimeter -> "%s"') % delimiter)
            row.append(x)
        wr.writerow(row)
    content.seek(0)  # Set index to 0, and start reading
    out_file = base64.b64encode(content.getvalue().encode('utf-8'))
    return out_file


def pos2idx(pos):
    match = re.match(r"([a-z]+)([0-9]+)", pos, re.I)
    if not match:
        raise ValidationError(_('Position %s is not valid') % (pos, ))
    col, row = match.groups()
    col_num = 0
    for c in col:
        if c in string.ascii_letters:
            col_num = col_num * 26 + (ord(c.upper()) - ord('A')) + 1
    return (int(row) - 1, col_num - 1)


def _get_cell_value(cell, field_type=False):
    """ If Odoo's field type is known, convert to valid string for import,
    if not know, just get value  as is """
    value = False
    datemode = 0  # From book.datemode, but we fix it for simplicity
    if field_type in ['date', 'datetime']:
        ctype = xlrd.sheet.ctype_text.get(cell.ctype, 'unknown type')
        if ctype == 'number':
            time_tuple = xlrd.xldate_as_tuple(cell.value, datemode)
            date = dt(*time_tuple)
            if field_type == 'date':
                value = date.strftime("%Y-%m-%d")
            elif field_type == 'datetime':
                value = date.strftime("%Y-%m-%d %H:%M:%S")
        else:
            value = cell.value
    elif field_type in ['integer', 'float']:
        value_str = str(cell.value).strip().replace(',', '')
        if len(value_str) == 0:
            value = ''
        elif value_str.replace('.', '', 1).isdigit():  # Is number
            if field_type == 'integer':
                value = int(float(value_str))
            elif field_type == 'float':
                value = float(value_str)
        else:  # Is string, no conversion
            value = value_str
    elif field_type in ['many2one']:
        # If number, change to string
        if isinstance(cell.value, (int, float, complex)):
            value = str(cell.value)
        else:
            value = cell.value
    else:  # text, char
        value = cell.value
    # If string, cleanup
    if isinstance(value, str):
        # value = value.encode('utf-8')
        if value[-2:] == '.0':
            value = value[:-2]
    # Except boolean, when no value, we should return as ''
    if field_type not in ['boolean']:
        if not value:
            value = ''
    return value


def _add_column(column_name, column_value, file_txt):
    i = 0
    txt_lines = []
    for line in file_txt.split('\n'):
        if line and i == 0:
            line = '"' + str(column_name) + '",' + line
        elif line:
            line = '"' + str(column_value) + '",' + line
        txt_lines.append(line)
        i += 1
    file_txt = '\n'.join(txt_lines)
    return file_txt


def _add_id_column(file_txt):
    i = 0
    txt_lines = []
    for line in file_txt.split('\n'):
        if line and i == 0:
            line = '"id",' + line
        elif line:
            line = '%s.%s' % ('xls', uuid.uuid4()) + ',' + line
        txt_lines.append(line)
        i += 1
    file_txt = '\n'.join(txt_lines)
    return file_txt
